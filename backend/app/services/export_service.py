from __future__ import annotations
import os
import csv
import uuid
from io import StringIO, BytesIO
from datetime import datetime
from typing import Optional, Any
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Color scheme
NAVY = "1B2A4A"
GOLD = "D4AF37"
LIGHT_GRAY = "F5F5F5"
WHITE = "FFFFFF"
DARK_GRAY = "333333"


def _navy_header_style(ws, row: int, col_count: int):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
        cell.font = Font(color=WHITE, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)


def build_csv(
    model_run: Any,
    fits: list[Any],
    top_n: int = 3,
) -> bytes:
    """Build CSV bytes with match, rank, predicted_score, ep, zero_prob, variance."""
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "match_id", "match_label", "rank",
        "predicted_home", "predicted_away",
        "predicted_score", "expected_points",
        "zero_point_probability", "variance",
        "fit_status",
    ])

    for fit in fits:
        match = getattr(fit, "match", None)
        if match:
            home = getattr(match.home_team, "name", None) or match.home_placeholder or "Home"
            away = getattr(match.away_team, "name", None) or match.away_placeholder or "Away"
            match_label = f"{home} vs {away}"
        else:
            match_label = str(fit.match_id)

        recs = sorted(fit.score_recommendations, key=lambda r: r.rank)[:top_n]
        for rec in recs:
            writer.writerow([
                str(fit.match_id),
                match_label,
                rec.rank,
                rec.predicted_home_goals,
                rec.predicted_away_goals,
                f"{rec.predicted_home_goals}-{rec.predicted_away_goals}",
                f"{float(rec.expected_points or 0):.4f}",
                f"{float(rec.zero_point_probability or 0):.4f}",
                f"{float(rec.variance_points or 0):.4f}",
                fit.fit_status or "",
            ])

    return output.getvalue().encode("utf-8")


def build_excel(
    model_run: Any,
    fits: list[Any],
    top_n: int = 3,
) -> bytes:
    """Build Excel bytes with multiple tabs."""
    wb = openpyxl.Workbook()

    # ── Tab 1: Recommended Picks ──────────────────────────────────────────
    ws_picks = wb.active
    ws_picks.title = "Recommended Picks"

    headers = [
        "Match", "Rank", "Score",
        "Home Goals", "Away Goals",
        "Expected Points", "Zero Prob", "Variance",
        "Score Probability",
    ]
    ws_picks.append(headers)
    _navy_header_style(ws_picks, 1, len(headers))

    for fit in fits:
        match = getattr(fit, "match", None)
        if match:
            home = getattr(match.home_team, "name", None) or match.home_placeholder or "Home"
            away = getattr(match.away_team, "name", None) or match.away_placeholder or "Away"
            match_label = f"{home} vs {away}"
        else:
            match_label = str(fit.match_id)

        recs = sorted(fit.score_recommendations, key=lambda r: r.rank)[:top_n]
        for i, rec in enumerate(recs):
            row_fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid") if i % 2 == 0 else None
            row = [
                match_label,
                rec.rank,
                f"{rec.predicted_home_goals}-{rec.predicted_away_goals}",
                rec.predicted_home_goals,
                rec.predicted_away_goals,
                round(float(rec.expected_points or 0), 4),
                round(float(rec.zero_point_probability or 0), 4),
                round(float(rec.variance_points or 0), 4),
                round(float(rec.score_probability or 0), 6),
            ]
            ws_picks.append(row)
            if row_fill:
                for col in range(1, len(row) + 1):
                    ws_picks.cell(row=ws_picks.max_row, column=col).fill = row_fill

    _auto_width(ws_picks)

    # ── Tab 2: Top Alternatives ───────────────────────────────────────────
    ws_alt = wb.create_sheet("Top Alternatives")
    alt_headers = ["Match", "Rank", "Score", "Expected Points", "Zero Prob", "Variance"]
    ws_alt.append(alt_headers)
    _navy_header_style(ws_alt, 1, len(alt_headers))

    for fit in fits:
        match = getattr(fit, "match", None)
        if match:
            home = getattr(match.home_team, "name", None) or match.home_placeholder or "Home"
            away = getattr(match.away_team, "name", None) or match.away_placeholder or "Away"
            match_label = f"{home} vs {away}"
        else:
            match_label = str(fit.match_id)

        recs = sorted(fit.score_recommendations, key=lambda r: r.rank)[top_n:top_n * 2]
        for rec in recs:
            ws_alt.append([
                match_label,
                rec.rank,
                f"{rec.predicted_home_goals}-{rec.predicted_away_goals}",
                round(float(rec.expected_points or 0), 4),
                round(float(rec.zero_point_probability or 0), 4),
                round(float(rec.variance_points or 0), 4),
            ])

    _auto_width(ws_alt)

    # ── Tab 3: Match Diagnostics ──────────────────────────────────────────
    ws_diag = wb.create_sheet("Match Diagnostics")
    diag_headers = [
        "Match", "Fit Status", "Lambda Home", "Lambda Away",
        "Total Goals", "RMSE",
        "Fitted Home Win", "Fitted Draw", "Fitted Away Win",
        "Market Home Win", "Market Draw", "Market Away Win",
        "Fit Error", "Warnings",
    ]
    ws_diag.append(diag_headers)
    _navy_header_style(ws_diag, 1, len(diag_headers))

    for fit in fits:
        match = getattr(fit, "match", None)
        if match:
            home = getattr(match.home_team, "name", None) or match.home_placeholder or "Home"
            away = getattr(match.away_team, "name", None) or match.away_placeholder or "Away"
            match_label = f"{home} vs {away}"
        else:
            match_label = str(fit.match_id)

        diag = fit.diagnostics or {}
        warnings = "; ".join(diag.get("warnings", []))
        total_goals = (float(fit.lambda_home or 0) + float(fit.lambda_away or 0))

        ws_diag.append([
            match_label,
            fit.fit_status or "",
            round(float(fit.lambda_home or 0), 6),
            round(float(fit.lambda_away or 0), 6),
            round(total_goals, 4),
            round(float(diag.get("rmse", 0)), 6),
            round(float(fit.fitted_home_win_prob or 0), 6),
            round(float(fit.fitted_draw_prob or 0), 6),
            round(float(fit.fitted_away_win_prob or 0), 6),
            round(float(fit.market_home_win_prob or 0), 6),
            round(float(fit.market_draw_prob or 0), 6),
            round(float(fit.market_away_win_prob or 0), 6),
            round(float(fit.fit_error or 0), 8),
            warnings,
        ])

    _auto_width(ws_diag)

    # ── Tab 4: Odds Consensus ─────────────────────────────────────────────
    ws_odds = wb.create_sheet("Odds Consensus")
    odds_headers = [
        "Match", "Home Win Prob", "Draw Prob", "Away Win Prob",
        "Over 2.5 Prob", "Under 2.5 Prob",
    ]
    ws_odds.append(odds_headers)
    _navy_header_style(ws_odds, 1, len(odds_headers))

    for fit in fits:
        match = getattr(fit, "match", None)
        if match:
            home = getattr(match.home_team, "name", None) or match.home_placeholder or "Home"
            away = getattr(match.away_team, "name", None) or match.away_placeholder or "Away"
            match_label = f"{home} vs {away}"
        else:
            match_label = str(fit.match_id)

        diag = fit.diagnostics or {}
        targets = diag.get("market_targets", {})

        ws_odds.append([
            match_label,
            round(float(fit.market_home_win_prob or targets.get("home_win", 0) or 0), 6),
            round(float(fit.market_draw_prob or targets.get("draw", 0) or 0), 6),
            round(float(fit.market_away_win_prob or targets.get("away_win", 0) or 0), 6),
            round(float(targets.get("over_2_5", 0) or 0), 6),
            round(float(targets.get("under_2_5", 0) or 0), 6),
        ])

    _auto_width(ws_odds)

    # ── Tab 5: Manual Overrides ───────────────────────────────────────────
    ws_over = wb.create_sheet("Manual Overrides")
    over_headers = ["Match ID", "Market Key", "Line", "Outcome Type", "Price Decimal", "Enabled", "Reason"]
    ws_over.append(over_headers)
    _navy_header_style(ws_over, 1, len(over_headers))

    for fit in fits:
        match = getattr(fit, "match", None)
        if match and hasattr(match, "manual_overrides"):
            for ov in match.manual_overrides:
                ws_over.append([
                    str(fit.match_id),
                    ov.market_key,
                    float(ov.line) if ov.line else None,
                    ov.outcome_type,
                    float(ov.price_decimal),
                    ov.enabled,
                    ov.reason or "",
                ])

    _auto_width(ws_over)

    # ── Tab 6: Scoring Rules ──────────────────────────────────────────────
    ws_rules = wb.create_sheet("Scoring Rules")
    rules_headers = ["Code", "Label", "Points", "Enabled", "Specificity Rank", "Description"]
    ws_rules.append(rules_headers)
    _navy_header_style(ws_rules, 1, len(rules_headers))

    pool_config = getattr(model_run, "pool_config", None)
    if pool_config and hasattr(pool_config, "scoring_rules"):
        for rule in sorted(pool_config.scoring_rules, key=lambda r: r.display_specificity_rank):
            ws_rules.append([
                rule.code,
                rule.label,
                float(rule.points),
                rule.enabled,
                rule.display_specificity_rank,
                rule.description or "",
            ])

    _auto_width(ws_rules)

    # ── Tab 7: Model Parameters ───────────────────────────────────────────
    ws_params = wb.create_sheet("Model Parameters")
    ws_params.append(["Parameter", "Value"])
    _navy_header_style(ws_params, 1, 2)

    params = model_run.parameters or {}
    for k, v in params.items():
        ws_params.append([k, str(v)])

    if pool_config:
        ws_params.append(["pool_config_name", pool_config.name])
        ws_params.append(["candidate_max_goals", pool_config.candidate_max_goals])
        ws_params.append(["ranking_metric", pool_config.ranking_metric])

    _auto_width(ws_params)

    # ── Tab 8: Run Metadata ───────────────────────────────────────────────
    ws_meta = wb.create_sheet("Run Metadata")
    ws_meta.append(["Field", "Value"])
    _navy_header_style(ws_meta, 1, 2)

    ws_meta.append(["run_id", str(model_run.id)])
    ws_meta.append(["run_type", model_run.run_type])
    ws_meta.append(["status", model_run.status])
    ws_meta.append(["started_at", str(model_run.started_at or "")])
    ws_meta.append(["completed_at", str(model_run.completed_at or "")])
    ws_meta.append(["matches_processed", len(fits)])
    ws_meta.append(["generated_at", datetime.utcnow().isoformat()])

    _auto_width(ws_meta)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def save_export(content: bytes, export_dir: str, filename: str) -> str:
    """Save export bytes to disk and return path."""
    os.makedirs(export_dir, exist_ok=True)
    path = os.path.join(export_dir, filename)
    with open(path, "wb") as f:
        f.write(content)
    return path
